from flask import Flask, render_template, request, redirect, url_for, send_from_directory, flash
import os, subprocess, uuid, shutil, sys
from pathlib import Path

app = Flask(__name__)
app.secret_key = "replace-with-a-secret-key"
BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_PDF = "stickers_msg_final.pdf"  # as your script writes by default

UPLOAD_DIR.mkdir(exist_ok=True)

# Path to the original script (unchanged)
ORIGINAL_SCRIPT = BASE_DIR / "sticker_generator_clean.py"


@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")


@app.route("/generate", methods=["POST"])
def generate():
    # 1) Validate uploaded file presence
    f = request.files.get("excel_file")
    if not f or f.filename == "":
        flash("Please upload an Excel (.xlsx) file.")
        return redirect(url_for("index"))

    # 2) Quick extension check
    allowed_ext = {".xlsx", ".xls"}
    ext = Path(f.filename).suffix.lower()
    if ext not in allowed_ext:
        flash("Invalid file type. Please upload an Excel file (.xlsx or .xls).")
        return redirect(url_for("index"))

    # 3) Validate Excel file content (try openpyxl)
    try:
        from openpyxl import load_workbook
    except Exception:
        flash("Server missing dependency: openpyxl. Please ask admin to install dependencies.")
        return redirect(url_for("index"))

    try:
        f.stream.seek(0)
        # load_workbook accepts a file-like; if it fails -> not real excel
        load_workbook(f, read_only=True)
        f.stream.seek(0)
    except Exception:
        flash("Invalid file content. Please upload a real Excel file (.xlsx or .xls).")
        return redirect(url_for("index"))

    # 4) Passed validation — save file to a unique workdir and run original script
    runid = uuid.uuid4().hex[:8]
    workdir = UPLOAD_DIR / runid
    workdir.mkdir(parents=True, exist_ok=True)

    saved_excel = workdir / f.filename
    f.save(saved_excel)

    # form options
    font_choice = request.form.get("font_choice", "1")  # "1","2","3"
    include_logo = request.form.get("include_logo", "n")  # 'y' or 'n'
    include_qr = request.form.get("include_qr", "n")      # 'y' or 'n'

    # Build stdin answers exactly as original script expects
    answers = "\n".join([
        font_choice,
        str(saved_excel),   # Excel file path
        "y" if include_logo == "y" else "n",
        "y" if include_qr == "y" else "n"
    ]) + "\n"

    # Run the original script as subprocess in the workdir
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUTF8"] = "1"

    cmd = [sys.executable, str(ORIGINAL_SCRIPT)]
    try:
        proc = subprocess.run(
            cmd,
            input=answers.encode("utf-8"),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            cwd=str(workdir),
            timeout=300,
            env=env,
            check=False
        )
    except subprocess.TimeoutExpired:
        flash("Processing timed out. The script took too long.")
        shutil.rmtree(workdir, ignore_errors=True)
        return redirect(url_for("index"))

    out = proc.stdout.decode("utf-8", errors="replace") if proc.stdout else ""
    err = proc.stderr.decode("utf-8", errors="replace") if proc.stderr else ""

    generated_pdf = workdir / OUTPUT_PDF
    project_pdf = BASE_DIR / OUTPUT_PDF
    # If script wrote PDF into project root (some scripts do), move it into workdir
    if not generated_pdf.exists() and project_pdf.exists():
        try:
            shutil.move(str(project_pdf), str(generated_pdf))
        except Exception as e:
            (workdir / "move_error.txt").write_text(str(e), encoding="utf-8")

    if not generated_pdf.exists():
        # save logs to the run folder for inspection
        (workdir / "script_stdout.txt").write_text(out, encoding="utf-8")
        (workdir / "script_stderr.txt").write_text(err, encoding="utf-8")
        # show result page with server output so you can debug
        return render_template("result.html", success=False, stdout=out, stderr=err, runid=runid)

    # Success
    return render_template(
        "result.html",
        success=True,
        runid=runid,
        pdf_name=generated_pdf.name
    )


@app.route("/download/<runid>/<filename>", methods=["GET"])
def download(runid, filename):
    folder = UPLOAD_DIR / runid
    if not folder.exists():
        flash("File not found.")
        return redirect(url_for("index"))
    return send_from_directory(str(folder), filename, as_attachment=True)


if __name__ == "__main__":
    # helpful debug: print available routes on startup so you can confirm 'download' exists
    print("URL map:\n", app.url_map)
    app.run(debug=True, port=5000)
